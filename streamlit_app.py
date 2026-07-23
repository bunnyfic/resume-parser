import os
import streamlit as st
import pandas as pd
from io import BytesIO
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- pipeline imports ---
from src.text_extractor.pdf_extractor import extract_text_from_pdf
from src.text_extractor.image_extractor import is_ocr_available
from src.preprocessing.text_cleaner import clean_text
from src.extraction.email_extractor import extract_email
from src.extraction.phone_extractor import extract_phone
from src.extraction.name_extractor import extract_name_from_pdf


SUPPORTED_PDF_EXTENSIONS = {".pdf"}

NEON_GREEN = "39FF14"
BG_DARK = "0E1117"
CARD_DARK = "1B1F27"
LIGHT_TEXT = "E6E6E6"


def process_resume(file_path: str) -> dict:
    ext = os.path.splitext(file_path)[1].lower()

    if ext not in SUPPORTED_PDF_EXTENSIONS:
        return None

    raw_text = extract_text_from_pdf(file_path)
    name = extract_name_from_pdf(file_path)
    cleaned_text = clean_text(raw_text)

    return {
        "Name": name,
        "Email": extract_email(cleaned_text),
        "Phone": extract_phone(cleaned_text),
        "File": os.path.basename(file_path)
    }


def build_styled_excel(df: pd.DataFrame) -> BytesIO:
    """Write df to an in-memory xlsx and apply a neon-green / dark theme."""
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl", sheet_name="Resumes")
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    header_fill = PatternFill(start_color=NEON_GREEN, end_color=NEON_GREEN, fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color=BG_DARK)
    header_align = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Calibri", size=11, color=BG_DARK)
    body_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color=NEON_GREEN),
        right=Side(style="thin", color=NEON_GREEN),
        top=Side(style="thin", color=NEON_GREEN),
        bottom=Side(style="thin", color=NEON_GREEN),
    )

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Body rows — light fill, dark readable text, thin neon borders
    body_fill = PatternFill(start_color="F5FFF0", end_color="F5FFF0", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border
            cell.fill = body_fill

    # Auto-size columns
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 6

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ---------------- UI ----------------
st.set_page_config(page_title="Resume Parser", page_icon="🟢", layout="wide")

# ---------- THEME ----------
st.markdown(
    f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@600;700&family=Inter:wght@400;500;600&display=swap');

        html, body, [class*="css"] {{
            font-family: 'Inter', sans-serif;
        }}

        .stApp {{
            background-color: #{BG_DARK};
        }}

        /* ---- Header ---- */
        .app-header {{
            padding: 1.6rem 0 1rem 0;
            border-bottom: 1px solid #2A2F3A;
            margin-bottom: 1.5rem;
        }}
        .app-header h1 {{
            font-family: 'Poppins', sans-serif;
            font-weight: 700;
            font-size: 2.4rem;
            color: #{LIGHT_TEXT};
            letter-spacing: 0.5px;
            margin-bottom: 0.2rem;
        }}
        .app-header h1 span {{
            color: #{NEON_GREEN};
            text-shadow: 0 0 12px rgba(57, 255, 20, 0.55);
        }}
        .app-header p {{
            font-family: 'Inter', sans-serif;
            color: #9AA0AC;
            font-size: 1rem;
            margin-top: 0;
        }}

        /* ---- Upload area ---- */
        section[data-testid="stFileUploaderDropzone"] {{
            background-color: #{CARD_DARK};
            border: 1.5px dashed #{NEON_GREEN}66;
            border-radius: 10px;
        }}
        section[data-testid="stFileUploaderDropzone"]:hover {{
            border: 1.5px dashed #{NEON_GREEN};
        }}

        /* ---- Dataframe ---- */
        div[data-testid="stDataFrame"] {{
            background-color: #{CARD_DARK};
            border: 1px solid #2A2F3A;
            border-radius: 8px;
            padding: 6px;
        }}
        thead tr th {{
            background-color: #23272F !important;
            color: #{LIGHT_TEXT} !important;
            font-family: 'Inter', sans-serif;
        }}

        /* ---- Section subheaders ---- */
        h2, h3 {{
            font-family: 'Poppins', sans-serif;
            color: #{LIGHT_TEXT} !important;
        }}

        /* ---- Download button ---- */
        .stDownloadButton button {{
            background-color: #{NEON_GREEN};
            color: #{BG_DARK};
            border-radius: 8px;
            border: none;
            padding: 0.7em 1.6em;
            font-family: 'Poppins', sans-serif;
            font-weight: 600;
            font-size: 0.95rem;
            box-shadow: 0 0 14px rgba(57, 255, 20, 0.45);
            transition: all 0.15s ease-in-out;
        }}
        .stDownloadButton button:hover {{
            background-color: #ffffff;
            color: #{BG_DARK};
            box-shadow: 0 0 22px rgba(57, 255, 20, 0.85);
            transform: translateY(-1px);
        }}

        /* ---- Progress bar ---- */
        div[data-testid="stProgress"] > div > div {{
            background-color: #{NEON_GREEN};
        }}

        /* ---- Alerts ---- */
        div[data-testid="stAlert"] {{
            border-radius: 8px;
            font-family: 'Inter', sans-serif;
        }}

        /* ---- Sidebar ---- */
        section[data-testid="stSidebar"] {{
            background-color: #{CARD_DARK};
            border-right: 1px solid #2A2F3A;
        }}
        section[data-testid="stSidebar"] h1,
        section[data-testid="stSidebar"] h2,
        section[data-testid="stSidebar"] h3 {{
            font-family: 'Poppins', sans-serif;
            color: #{LIGHT_TEXT} !important;
        }}
        section[data-testid="stSidebar"] label {{
            font-family: 'Inter', sans-serif;
            color: #{LIGHT_TEXT} !important;
        }}

        /* ---- Neon info boxes (About page) ---- */
        .neon-box {{
            background-color: #{CARD_DARK};
            border: 1px solid #{NEON_GREEN}55;
            border-left: 4px solid #{NEON_GREEN};
            border-radius: 10px;
            padding: 1.3rem 1.6rem;
            margin-bottom: 1.4rem;
            box-shadow: 0 0 18px rgba(57, 255, 20, 0.08);
        }}
        .neon-box h3 {{
            font-family: 'Poppins', sans-serif;
            color: #{NEON_GREEN} !important;
            font-size: 1.15rem;
            margin-top: 0;
            margin-bottom: 0.7rem;
            text-shadow: 0 0 10px rgba(57, 255, 20, 0.35);
        }}
        .neon-box p, .neon-box li {{
            font-family: 'Inter', sans-serif;
            color: #{LIGHT_TEXT};
            font-size: 0.96rem;
            line-height: 1.55;
        }}
        .neon-box ul {{
            margin: 0;
            padding-left: 1.2rem;
        }}
        .neon-box b, .neon-box strong {{
            color: #{NEON_GREEN};
        }}
        .neon-box code {{
            background-color: #{BG_DARK};
            color: #{NEON_GREEN};
            border-radius: 4px;
            padding: 0.1rem 0.35rem;
            font-size: 0.88rem;
        }}
        .neon-divider {{
            border: none;
            height: 1px;
            background: linear-gradient(90deg, transparent, #{NEON_GREEN}88, transparent);
            margin: 1.8rem 0;
        }}
    </style>
    """,
    unsafe_allow_html=True
)

# ---------- SIDEBAR NAVIGATION ----------
st.sidebar.markdown(
    f"""
    <h2 style="color:#{NEON_GREEN}; text-shadow:0 0 10px rgba(57,255,20,0.4);
    font-family:'Poppins', sans-serif; margin-bottom:0.2rem;">📋 Menu</h2>
    """,
    unsafe_allow_html=True
)
page = st.sidebar.radio(
    "Navigate",
    ["🏠 Resume Parser", "ℹ️ About the App"],
    label_visibility="collapsed"
)


def render_about_page():
    st.markdown(
        """
        <div class="app-header">
            <h1>About <span>the App</span></h1>
            <p>How resume parsing works, why ATS formatting matters, and the technology behind it.</p>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown(
        f"""
        <div class="neon-box">
            <h3>📄 How Recruiters Extract Name, Email &amp; Phone from ATS-Compatible PDFs</h3>
            <p>
            An <b>ATS-compatible (Applicant Tracking System-compatible) resume</b> is built as a
            genuine <b>text-layer PDF</b> — meaning every character, from the candidate's name to
            their contact details, is stored as selectable, machine-readable text rather than a
            picture. This is what makes automated extraction reliable:
            </p>
            <ul>
                <li><b>Text extraction:</b> Tools like <code>PyPDF2</code>, <code>pdfplumber</code>,
                or <code>PyMuPDF</code> read the embedded text layer directly from the PDF's internal
                structure, preserving reading order line by line.</li>
                <li><b>Name detection:</b> The candidate's name is usually the largest / topmost text
                block on page one, so parsers combine positional heuristics (font size, position) with
                NLP-based Named Entity Recognition (e.g. spaCy's <code>PERSON</code> entity) to isolate it
                confidently.</li>
                <li><b>Email extraction:</b> A regular expression such as
                <code>[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\\.[a-zA-Z0-9-.]+</code> reliably matches the
                standard email format, wherever it appears in the cleaned text.</li>
                <li><b>Phone extraction:</b> Regex patterns tuned for international and local formats
                (with optional country codes, brackets, and separators) pull out digit sequences that
                match valid phone-number shapes.</li>
            </ul>
            <p>
            Because everything sits in a single, linear text stream, this pipeline works
            <b>fast and near-perfectly</b> on well-structured, single-column resumes.
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown(
        f"""
        <div class="neon-box">
            <h3>⚠️ Common Flaws in Non-ATS-Friendly Resumes</h3>
            <p>
            Many visually appealing resume templates (often built in <b>Canva, Photoshop, or heavily
            designed Word/InDesign templates</b>) actively break automated parsing. Typical issues include:
            </p>
            <ul>
                <li><b>Multi-column layouts:</b> Parsers read left-to-right, top-to-bottom in the order
                text is stored in the PDF — not the order it visually appears. A two-column resume can
                interleave the "Skills" sidebar with the "Experience" section mid-sentence.</li>
                <li><b>Text inside tables or text boxes:</b> Content placed in graphic text boxes is
                sometimes stored out of sequence, or excluded from the extractable text layer entirely.</li>
                <li><b>Contact info as images/icons:</b> Phone and email icons paired with text rendered
                as part of a graphic (rather than real text) are invisible to text extractors.</li>
                <li><b>Headers &amp; footers:</b> Many ATS parsers ignore header/footer regions — a
                candidate's name or contact details placed there can be silently dropped.</li>
                <li><b>Scanned or "flattened" PDFs:</b> A resume exported as a flattened image, or a
                phone-scanned photo saved as PDF, has <b>no text layer at all</b> — only pixels — so
                traditional extraction returns nothing.</li>
                <li><b>Unusual fonts &amp; special characters:</b> Decorative or embedded fonts can be
                mis-encoded during extraction, turning readable text into garbled symbols.</li>
                <li><b>Non-standard section headings:</b> Creative headings like "My Journey" instead of
                "Experience" can confuse rule-based section detection used in deeper ATS scoring.</li>
            </ul>
            <p>
            The safest resumes for ATS parsing are <b>single-column, plain-text-based PDFs with
            standard fonts and conventional section headers</b> — visually simple, but far more reliably read.
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown(
        f"""
        <div class="neon-box">
            <h3>⚙️ Tech Stack Behind This Parser</h3>
            <p>This app uses a two-tier extraction pipeline so it can handle both clean, text-based
            PDFs and problematic image-based ones:</p>
            <ul>
                <li><b>Native text extraction:</b> For standard, text-layer PDFs, the parser reads the
                embedded text directly — fast, accurate, and requiring no image processing.</li>
                <li><b>OCR fallback (Tesseract OCR):</b> When a PDF has little or no extractable text
                — for example, a <b>scanned resume</b>, a <b>photo saved as PDF</b>, or a
                <b>Word/Docs file "printed" to PDF where text was flattened into an image</b> — the
                app falls back to <code>pytesseract</code> (a Python wrapper around Google's Tesseract
                OCR engine). It rasterizes each PDF page into an image, then runs optical character
                recognition to reconstruct the readable text before running it through the same
                name/email/phone extraction logic.</li>
                <li><b>Text cleaning:</b> A preprocessing layer normalizes whitespace, line breaks, and
                encoding artifacts (common after OCR) before extraction rules are applied, improving
                match accuracy for both regex and NER-based extraction.</li>
                <li><b>Regex &amp; rule-based extraction:</b> Dedicated modules handle email and phone
                number extraction using tuned regular expressions, while name extraction combines
                layout/positional cues with the PDF metadata and text structure.</li>
                <li><b>Streamlit:</b> Powers the interactive front-end — file uploads, progress
                tracking, live status updates, and the results table.</li>
                <li><b>Pandas + OpenPyXL:</b> Structured results are compiled into a DataFrame and
                exported to a styled <code>.xlsx</code> file with a neon-green header, borders, and
                auto-sized columns for a clean, recruiter-ready spreadsheet.</li>
            </ul>
            <p>
            In short: <b>clean text PDFs are read directly</b>, while <b>image-embedded or scanned
            PDFs are recovered via OCR</b> — giving the app broad coverage across both ATS-friendly
            and non-ATS-friendly resume formats, even though OCR results are naturally less precise
            than native text extraction.
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )


def render_parser_page():
    st.markdown(
        """
        <div class="app-header">
            <h1>Resume <span>Parser</span></h1>
            <p>Upload PDF resumes and export clean, structured candidate data to Excel.</p>
        </div>
        """,
        unsafe_allow_html=True
    )

    # ---------- ENVIRONMENT CHECK ----------
    if not is_ocr_available():
        st.warning(
            "⚠️ OCR engine (Tesseract) not detected on this server. Scanned "
            "PDFs won't be read correctly — only text-based PDFs will work. "
            "If this app is deployed on Streamlit Cloud, add a `packages.txt` "
            "file to your repo root containing the line `tesseract-ocr` and redeploy."
        )

    # ---------- SESSION STATE ----------
    if "results" not in st.session_state:
        st.session_state.results = []

    if "processed_files" not in st.session_state:
        st.session_state.processed_files = set()

    uploaded_files = st.file_uploader(
        "Upload resumes (PDF only)",
        type=["pdf"],
        accept_multiple_files=True
    )

    # ---------- PROCESS FILES ----------
    if uploaded_files:
        new_files = [
            f for f in uploaded_files
            if f.name not in st.session_state.processed_files
        ]

        if new_files:
            progress = st.progress(0)
            status = st.empty()
            failed_files = []

            with tempfile.TemporaryDirectory() as temp_dir:
                for i, file in enumerate(new_files, start=1):
                    status.text(f"Processing: {file.name}")

                    temp_path = os.path.join(temp_dir, file.name)
                    with open(temp_path, "wb") as f:
                        f.write(file.read())

                    result = process_resume(temp_path)
                    if result:
                        st.session_state.results.append(result)
                    else:
                        failed_files.append(file.name)

                    st.session_state.processed_files.add(file.name)
                    progress.progress(i / len(new_files))

            status.empty()
            progress.empty()

            if failed_files:
                st.error(f"❌ Could not process {len(failed_files)} file(s): {', '.join(failed_files)}")
            if len(new_files) - len(failed_files) > 0:
                st.success(f"✅ Successfully processed {len(new_files) - len(failed_files)} file(s)!")

    # ---------- DISPLAY ----------
    if st.session_state.results:
        df = pd.DataFrame(st.session_state.results)

        st.subheader("Extracted Data")
        st.dataframe(df, use_container_width=True)

        excel_buffer = build_styled_excel(df)

        st.download_button(
            label="⬇  Download Excel",
            data=excel_buffer,
            file_name="resume_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# ---------- PAGE ROUTING ----------
if page == "ℹ️ About the App":
    render_about_page()
else:
    render_parser_page()
